import os
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import json
import glob
import shutil
import datetime
from openpyxl import load_workbook
from scripts.excel_utils import get_latest_inventory_sheet, get_taxonomy

def export_to_json(out_path="webapp/data/inventory.json", xlsx_id=None, pdf_id=None):
    print("Extracting latest inventory for the Web App...")

    # ── Read publish config ──────────────────────────────────────────
    publish_pricing = True
    config_path = os.path.join("webapp", "data", "publish_config.json")
    if os.path.exists(config_path):
        try:
            with open(config_path, "r") as cf:
                publish_pricing = json.load(cf).get("publish_pricing", True)
        except Exception:
            pass

    # ── CLEANUP webapp/data ──────────────────────────────────────────
    data_dir = os.path.dirname(out_path)
    if os.path.exists(data_dir):
        for f in os.listdir(data_dir):
            if f.endswith(".xlsx") or f.endswith(".pdf"):
                try:
                    os.remove(os.path.join(data_dir, f))
                except Exception:
                    pass

    # ── LOAD DATA ────────────────────────────────────────────────────
    year = datetime.datetime.now().year
    latest_out = os.path.join("output", f"{year} Golden Inventory.xlsx")
    if not os.path.exists(latest_out):
        print(f"Master workbook {latest_out} not found.")
        return
    
    wb = load_workbook(latest_out, data_only=True)
    sheet_name, recent_date = get_latest_inventory_sheet(wb)

    if not sheet_name or sheet_name not in wb.sheetnames:
        print("Could not find a valid data sheet.")
        wb.close()
        return
        
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    header_idx = -1
    for i, row in enumerate(rows):
        if any(str(cell).lower().strip() in ["item", "internal reference"] for cell in row if cell):
            header_idx = i
            break
            
    headers = [str(cell).strip().lower() if cell else "" for cell in rows[header_idx]]
    
    # ── Map Columns ──────────────────────────────────────────────────
    col_map = {"item": -1, "description": -1, "category": -1, "raw_category": -1, "price": -1, "quantity": -1, "condition": -1}
    
    for i, h in enumerate(headers):
        if h == "item" or "internal reference" in h: col_map["item"] = i
        elif h == "description" or "name" in h: col_map["description"] = i
        elif h == "category": col_map["category"] = i
        elif "product category" in h or "raw" in h: col_map["raw_category"] = i
        elif "price" in h: col_map["price"] = i
        elif "qty" in h or "available" in h or "quantity" in h or "on hand" in h: col_map["quantity"] = i
        elif "condition" in h or "status" in h: col_map["condition"] = i

    # ── Process Items ────────────────────────────────────────────────
    inventory = []
    for row in rows[header_idx+1:]:
        item_code = str(row[col_map["item"]]).strip() if col_map["item"] != -1 and row[col_map["item"]] else None
        if not item_code: continue
            
        desc = (str(row[col_map["description"]]) if col_map["description"] != -1 and row[col_map["description"]] else "Unknown Item")
        cat_raw = str(row[col_map["raw_category"]]) if col_map["raw_category"] != -1 and row[col_map["raw_category"]] else ""
        condition = str(row[col_map["condition"]]) if col_map["condition"] != -1 and row[col_map["condition"]] else ""
        
        price = row[col_map["price"]] if col_map["price"] != -1 else 0
        raw_avail = row[col_map["quantity"]] if col_map["quantity"] != -1 else 0
        avail = int(float(str(raw_avail))) if str(raw_avail).replace('.', '', 1).isdigit() else 0
            
        # Taxonomy lookup with condition
        tax = get_taxonomy(cat_raw, item_code, desc, condition=condition)
        
        # Image handling
        item_code_img = item_code.replace(" / ", " - ").replace("/", "-")
        web_img_path = f"images/{item_code_img}.png"
        if not os.path.exists(os.path.join("images", f"{item_code_img}.png")) and not os.path.exists(os.path.join("images", f"{item_code_img}.jpg")):
            web_img_path = "https://via.placeholder.com/150/f0f0f0/888888?text=No+Image"
            
        inventory.append({
            "id": item_code,
            "name": desc,
            "category": tax["parent"] if tax["parent"] == "Major Appliances" else tax["sub"],
            "parent_category": tax["parent"],
            "sub_category": tax["sub"],
            "price": f"{price:.2f}" if publish_pricing else "Hidden",
            "available": avail if publish_pricing else "Hidden",
            "image": web_img_path
        })
        
    wb.close()
    
    # ── Save JSON ────────────────────────────────────────────────────
    payload = {
        "last_updated": recent_date.strftime("%B %d, %Y"),
        "publish_pricing": publish_pricing,
        "items": inventory
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=4)
        
    print(f"Exported {len(inventory)} items to {out_path}.")

if __name__ == "__main__":
    export_to_json()
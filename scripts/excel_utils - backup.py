"""
excel_utils.py
Utility functions for Golden Inventory automation.
"""

import re
from datetime import datetime
from openpyxl import load_workbook
from scripts.item_overrides import ITEM_CATEGORY_OVERRIDES

# ─── Master 8-Pillar Taxonomy Rules ─────────────────────────────────────────
# ─── SKU PATTERN RULES ────────────────────────────────────────────────

SKU_PATTERN_RULES = [

    # Kitchen Appliances
    ("ED-", "Kitchen Appliances", "Heating & Cooking"),
    ("TS-", "Kitchen Appliances", "Air Fryer & Toaster"),
    ("FP",  "Kitchen Appliances", "Food Preparation"),
    ("MG-", "Kitchen Appliances", "Food Preparation"),
    ("CB",  "Kitchen Appliances", "Coffee & Tea"),

    # Audio
    ("J-",      "Audio", "Home Audio"),
    ("BTSPK",   "Audio", "Home Audio"),
    ("PBX-",    "Audio", "Home Audio"),
    ("SP-",     "Audio", "Home Audio"),

    # Electronics
    ("SM-STV", "Electronics", "TV Accessories"),

    # Air
    ("MPI-", "Air & Climate Control", "Fans"),
    ("LSL",  "Air & Climate Control", "Fans"),
]

MASTER_TAXONOMY_RULES = {

    "Major Appliances": {

        "Refrigerators": [
            "refrigerator", "fridge", "top freezer", "bottom freezer",
            "french door", "side by side", "counter depth"
        ],

        "Compact Refrigerators": [
            "compact refrigerator", "compact fridge",
            "mini fridge", "dorm fridge"
        ],

        "Beverage Centers": [
            "beverage center", "beverage cooler",
            "drink cooler"
        ],

        "Wine Coolers": [
            "wine cooler", "wine cellar"
        ],

        "Chest Freezers": [
            "chest freezer"
        ],

        "Upright Freezers": [
            "upright freezer",
            "convertible freezer"
        ],

        "Laundry": [
            "washer", "washing machine",
            "dryer", "stacked laundry",
            "laundry center"
        ],

        "Dishwashers": [
            "dishwasher"
        ],

        "Ranges": [
            "range", "gas range",
            "electric range"
        ],

        "Wall Ovens": [
            "wall oven"
        ],

        "Cooktops": [
            "cooktop", "induction cooktop"
        ],

        "Range Hoods": [
            "range hood", "hood"
        ]

    },

    "Kitchen Appliances": {

        "Coffee & Tea": [
            "coffee",
            "espresso",
            "keurig",
            "k-cup",
            "tea",
            "kettle",
            "coffee maker",
            "percolator",
            "urn",
            "milk frother"
        ],

        "Food Preparation": [
            "blender",
            "personal blender",
            "food processor",
            "processor",
            "chopper",
            "stand mixer",
            "hand mixer",
            "immersion blender",
            "hand blender",
            "juicer",
            "slow juicer",
            "cold press",
            "food chopper"
        ],

        "Air Fryer & Toaster": [
            "air fryer",
            "toaster",
            "toaster oven",
            "deep fryer"
        ],

        "Heating & Cooking": [
            "microwave",
            "hot plate",
            "hotplate",
            "burner",
            "double burner",
            "griddle",
            "pizza maker",
            "waffle",
            "waffle maker",
            "panini",
            "sandwich maker",
            "tortilla",
            "arepa",
            "dehydrator"
        ],

        "Pressure Cookers": [
            "pressure cooker",
            "pressure canner"
        ],

        "Rice Cookers": [
            "rice cooker"
        ],

        "Slow Cookers": [
            "slow cooker",
            "crock pot"
        ],

        "Cookware": [
            "cookware",
            "stock pot",
            "fry pan",
            "sauce pan",
            "skillet"
        ],

        "Kitchen Tools": [
            "knife sharpener",
            "can opener",
            "kitchen scale"
        ]

    },

    "Air & Climate Control": {

        "Portable AC": [
            "portable air conditioner",
            "portable ac"
        ],

        "Window AC": [
            "window air conditioner",
            "window ac"
        ],

        "Fans": [
            "fan",
            "tower fan",
            "pedestal fan",
            "desk fan",
            "neck fan",
            "box fan",
            "floor fan"
        ],

        "Air Purifiers": [
            "air purifier"
        ],

        "Humidifiers": [
            "humidifier"
        ],

        "Dehumidifiers": [
            "dehumidifier"
        ],

        "Heaters": [
            "heater",
            "ceramic heater",
            "radiator heater",
            "oil filled",
            "tower heater"
        ]

    },

    "Audio": {

        "Portable Audio": [
            "boombox",
            "cassette",
            "cd player",
            "portable cd",
            "walkman",
            "radio"
        ],

        "Home Audio": [
            "speaker",
            "bluetooth speaker",
            "soundbar",
            "subwoofer"
        ],

        "Headphones & Earphones": [
            "headphones",
            "earbuds",
            "earphones",
            "headset"
        ]

    },

    "Electronics": {

        "TV Accessories": [
            "antenna",
            "converter box"
        ],

        "TV Mounts": [
            "wall mount",
            "tv mount",
            "mount"
        ],

        "Remote Controls": [
            "remote",
            "universal remote"
        ],

        "Projectors": [
            "projector"
        ],

        "Media Players": [
            "media player",
            "dvd player",
            "streaming"
        ],

        "Security": [
            "security camera",
            "doorbell camera",
            "surveillance"
        ],

        "Car Electronics": [
            "car stereo",
            "fm transmitter"
        ],

        "Cables": [
            "hdmi",
            "usb cable",
            "audio cable"
        ],

        "Power Accessories": [
            "surge protector",
            "power strip",
            "extension cord"
        ]

    },

    "Health & Personal Care": {

        "Hair Care": [
            "hair dryer",
            "curling",
            "flat iron",
            "hot comb",
            "styler"
        ],

        "Personal Care": [
            "shaver",
            "trimmer",
            "clipper",
            "massager"
        ]

    },

    "Home & Office": {

        "Irons": [
            "iron",
            "steam iron"
        ],

        "Scales": [
            "scale",
            "body analysis"
        ],

        "Calculators": [
            "calculator"
        ],

        "Lighting": [
            "led light",
            "flashlight",
            "lamp"
        ],

        "Vacuums": [
            "vacuum"
        ],

        "Alarm Clocks": [
            "alarm clock"
        ]

    }

}


# ─── OVERRIDE → PARENT MAPPING ───────────────────────────────────────────────
# Maps override sub-category names (from item_overrides.py) that do NOT exactly
# match a MASTER_TAXONOMY_RULES key to their correct parent category.
# This ensures every string override resolves to a proper (parent, sub) pair.

OVERRIDE_PARENT_MAP = {
    # Major Appliances
    "Dishwasher":             "Major Appliances",
    "Coolers":                "Major Appliances",
    "Beverage Cooler":        "Major Appliances",
    "Wine Cooler":            "Major Appliances",
    "Refrigerators":          "Major Appliances",

    # Kitchen Appliances
    "Blenders & Mixers":      "Kitchen Appliances",
    "Electric Skillets":      "Kitchen Appliances",
    "Microwave":              "Kitchen Appliances",
    "Coffee Maker":           "Kitchen Appliances",
    "Toaster":                "Kitchen Appliances",
    "Rice Cooker":            "Kitchen Appliances",
    "Pressure Cooker":        "Kitchen Appliances",
    "Outdoor Cooking":        "Kitchen Appliances",
    "Kitchen accessories":    "Kitchen Appliances",

    # Air & Climate Control
    "Air Conditioner Accessories": "Air & Climate Control",

    # Audio
    "Speakers":               "Audio",
    "Soundbars":              "Audio",
    "Microphones":            "Audio",
    "Radios":                 "Audio",
    "Home Radio":             "Audio",
    "Boombox":                "Audio",
    "Audio Players":          "Audio",
    "Home Audio":             "Audio",

    # Electronics
    "TV Remote":              "Electronics",
    "TV Wall Bracket":        "Electronics",
    "CD/DVD Player":          "Electronics",
    "Computer Accessories":   "Electronics",
    "Phones":                 "Electronics",
    "Power Banks":            "Electronics",
    "Power Bank":             "Electronics",
    "Surge Protectors":       "Electronics",
    "Security Cameras":       "Electronics",
    "Car Audio":              "Electronics",
    "Phone Charger":          "Electronics",
    "Cables":                 "Electronics",

    # Health & Personal Care
    "Shavers":                "Health & Personal Care",
    "Trimmers":               "Health & Personal Care",

    # Home & Office
    "Irons":                  "Home & Office",
    "Timers":                 "Home & Office",
    "Steam Mop":              "Home & Office",
    "Lighting":               "Home & Office",
    "Calculator":             "Home & Office",

    # Catch-all for intentionally vague overrides
    "Accessories":            "Other",
    "Car Electronics":        "Electronics",
}


# ─── SKU NORMALIZATION ────────────────────────────────────────────────

def normalize_sku(sku: str) -> str:
    if not sku:
        return ""

    return (
        str(sku)
        .strip()
        .upper()
        .replace(" ", "")
        .replace("/", "-")
        .replace("--", "-")
    )

NORMALIZED_ITEM_CATEGORY_OVERRIDES = {normalize_sku(k): v for k, v in ITEM_CATEGORY_OVERRIDES.items()}


# ─── TAXONOMY ENGINE ────────────────────────────────────────────────

def get_taxonomy(product_category_path: str, item_id: str = None, description: str = ""):
    """
    Classify a product into (parent, sub) following this strict priority:
      1. Normalize SKU
      2. Explicit item override  (ITEM_CATEGORY_OVERRIDES)
      3. SKU prefix rules        (SKU_PATTERN_RULES)
      4. Keyword scoring         (MASTER_TAXONOMY_RULES)
      5. Refurbished tagging
      6. Fallback: Other / Uncategorized
    """
    sku = normalize_sku(item_id)

    # ── Step 2: Explicit item override (highest priority) ────────────────────
    if sku in NORMALIZED_ITEM_CATEGORY_OVERRIDES:
        override = NORMALIZED_ITEM_CATEGORY_OVERRIDES[sku]

        if isinstance(override, dict):
            return override

        if isinstance(override, str):
            # If the override itself is a valid parent category
            if override in MASTER_TAXONOMY_RULES:
                return {"parent": override, "sub": "Uncategorized"}
            # Try direct match against a MASTER_TAXONOMY_RULES sub-category key
            for parent, subs in MASTER_TAXONOMY_RULES.items():
                if override in subs:
                    return {"parent": parent, "sub": override}
            # Fall back to the explicit parent mapping for non-standard names
            if override in OVERRIDE_PARENT_MAP:
                return {"parent": OVERRIDE_PARENT_MAP[override], "sub": override}
            # Last resort: preserve the sub-category name under 'Other'
            return {"parent": "Other", "sub": override}

        return override  # Unknown override type: return as-is

    # ── Determine refurbished status (used in steps 3–6) ─────────────────────
    is_refurbished = sku.endswith(("/RBO", "/RB", "/ROA", "-RBO", "-RB", "-ROA"))

    # ── Step 3: SKU prefix rules ──────────────────────────────────────────────
    for prefix, parent, sub in SKU_PATTERN_RULES:
        # Normalize prefix the same way normalize_sku normalizes the SKU
        normalized_prefix = prefix.upper().replace(" ", "").replace("/", "-").replace("--", "-")
        if sku.startswith(normalized_prefix):
            if is_refurbished:
                return {"parent": "Refurbished", "sub": sub}
            return {"parent": parent, "sub": sub}

    # ── Step 4 & 5: Keyword scoring across ALL parent categories ─────────────
    test_str = " ".join([
        str(product_category_path or ""),
        str(description or ""),
        str(item_id or "")
    ]).lower()

    best_parent = None
    best_sub = None
    best_score = 0

    for parent, subcategories in MASTER_TAXONOMY_RULES.items():
        for sub, keywords in subcategories.items():
            score = 0
            for kw in keywords:
                kw_lower = kw.lower()
                if kw_lower in test_str:
                    score += len(kw_lower.split()) * 20
                for word in kw_lower.split():
                    if len(word) >= 3 and word in test_str:
                        score += 5
            if score > best_score:
                best_score = score
                best_parent = parent
                best_sub = sub
    # NOTE: return statements are OUTSIDE the outer loop so ALL categories are scored.

    # ── Step 6: Return best match or fallback ────────────────────────────────
    if best_parent:
        if is_refurbished:
            return {"parent": "Refurbished", "sub": best_sub}
        return {"parent": best_parent, "sub": best_sub}

    if is_refurbished:
        return {"parent": "Refurbished", "sub": "Uncategorized"}

    return {"parent": "Other", "sub": "Uncategorized"}




# ─── RAW DATA LOADER ────────────────────────────────────────────────

def load_raw_data(inventory_tool_path: str) -> list[dict]:
    wb = load_workbook(inventory_tool_path, data_only=True)

    if "RAW" not in wb.sheetnames:
        raise ValueError("RAW sheet missing")

    ws = wb["RAW"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() for h in rows[0]]

    data = []
    for row in rows[1:]:
        row_dict = dict(zip(headers, row))
        if any(v for v in row_dict.values() if v is not None):
            data.append(row_dict)

    wb.close()
    return data


# ─── CATEGORY WRAPPER ────────────────────────────────────────────────

def classify_category(product_category_path: str, item_id: str = None, description: str = "") -> str:
    return get_taxonomy(product_category_path, item_id, description)["sub"]

def get_latest_inventory_sheet(workbook):
    """
    Returns the newest timestamped inventory sheet and its date.
    """

    latest_sheet = None
    latest_datetime = None

    for sheet in workbook.sheetnames:

        if sheet == "Item Listing and Pricing":
            continue

        try:
            sheet_dt = datetime.strptime(sheet, "%b %d %Y - %I-%M %p")

            if latest_datetime is None or sheet_dt > latest_datetime:
                latest_datetime = sheet_dt
                latest_sheet = sheet

        except ValueError:
            continue

    if latest_sheet is None:
        raise Exception("No timestamped inventory sheet found.")

    return latest_sheet, latest_datetime.date()

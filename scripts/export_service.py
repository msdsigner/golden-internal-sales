"""
export_service.py
Creates the secondary "Golden Inventory March 25 2026.xlsx" report
and exports it to PDF using win32com (Excel automation on Windows).

Layout (per image #2):
  Row 1:     Date header, merged A1:D1
  Row 2–3:   Banner image, merged A2:D3
  Row 4:     Column headers: IMAGE | Item | Description | Category
  Row 5+:    Data rows with images in col A
"""

import os
import shutil
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

INPUT_DIR  = "input"
OUTPUT_DIR = "output"
IMAGES_DIR = "images"
BANNER_PATH = os.path.join(IMAGES_DIR, "banner.png")


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def create_report_and_pdf(gi_path: str, sheet_name: str, sheet_date: date, year: int, data_map: dict = None):
    """
    Reads cols A-D (Image, Item, Description, Category) from
    'Item Listing and Pricing' in the Golden Inventory workbook,
    creates a dated report xlsx, and exports to PDF.
    - If data_map is provided (item_code -> {desc, cat}), it writes values instead of formulas.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Build output file name using the run timestamp
    date_label = sheet_date.strftime("%b %d %Y - %I-%M %p").replace(" 0", " ")

    out_name = f"Golden Inventory - {date_label}"

    out_xlsx = os.path.join(OUTPUT_DIR, f"{out_name}.xlsx")
    out_pdf = os.path.join(OUTPUT_DIR, f"{out_name}.pdf")

    print(f"  Creating report: {out_xlsx}")

    # Load source workbook
    src_wb = load_workbook(gi_path)
    ws_src = src_wb["Item Listing and Pricing"]

    # Read item codes and formulas (or values) from rows 4+
    rows_data = []
    for row in ws_src.iter_rows(min_row=4, min_col=2, max_col=4, values_only=True):
        item_code = row[0]
        if item_code:
            desc = row[1]
            cat = row[2]
            # If map is provided, resolve formulas to values
            if data_map and str(item_code) in data_map:
                res = data_map[str(item_code)]
                desc = res.get('desc', desc)
                cat = res.get('cat', cat)
            
            rows_data.append((item_code, desc, cat))

    src_wb.close()    # Sort rows_data by category before starting report workbook
    rows_data.sort(key=lambda x: (x[2], x[0]))

    # ── Create report workbook ──
    rpt_wb = openpyxl.Workbook()
    rpt_ws = rpt_wb.active
    rpt_ws.title = "Golden Inventory"

    # Column widths (Description wider)
    rpt_ws.column_dimensions["A"].width = 20
    rpt_ws.column_dimensions["B"].width = 12
    rpt_ws.column_dimensions["C"].width = 38
    rpt_ws.column_dimensions["D"].width = 15

    # Points conversion: 1px ≈ 0.75pt
    RP_DATE_HEIGHT   = 22     # Increased for balance
    RP_BANNER_HEIGHT = 120    # 160px
    RP_ITEM_HEIGHT   = 105    # 140px

    # ── Row 1: Date header ──
    day_name = sheet_date.strftime("%A")
    month_day_year = sheet_date.strftime("%B %d, %Y").replace(" 0", " ")
    date_header = f"{day_name}, {month_day_year}"

    rpt_ws.merge_cells("A1:D1")
    cell_date = rpt_ws["A1"]
    cell_date.value = date_header
    cell_date.font  = Font(bold=True, size=13, color="1F497D")
    cell_date.alignment = Alignment(horizontal="center", vertical="center")
    rpt_ws.row_dimensions[1].height = RP_DATE_HEIGHT

    # ── Row 2: Banner ──
    rpt_ws.row_dimensions[2].height = RP_BANNER_HEIGHT
    rpt_ws.merge_cells("A2:D2")

    if os.path.exists(BANNER_PATH):
        try:
            banner = XLImage(BANNER_PATH)
            banner.width  = 680
            banner.height = 160
            # Anchor to A2:D2 in report (col 0-3, to 4)
            _from = AnchorMarker(col=0, row=1)
            to    = AnchorMarker(col=4, row=2)
            banner.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=to)
            rpt_ws.add_image(banner)
        except Exception as e:
            print(f"  Warning: could not add banner: {e}")

    # ── Row 3: Column headers ──
    headers = ["IMAGE", "Item", "Description", "Category"]
    fills   = ["D9D9D9", "D9D9D9", "D9D9D9", "FFC000"]
    for col_idx, (hdr, fill) in enumerate(zip(headers, fills), start=1):
        cell = rpt_ws.cell(row=3, column=col_idx, value=hdr)
        cell.font      = Font(bold=True)
        cell.fill      = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
    rpt_ws.row_dimensions[3].height = 20

    rpt_ws.column_dimensions["D"].width = 14

    # Row 3 Headings (Update column D heading alignment too)
    rpt_ws.cell(row=3, column=4).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Row 4+: Data ──
    for r_offset, (item_code, desc_val, cat_val) in enumerate(rows_data):
        r = r_offset + 4
        rpt_ws.row_dimensions[r].height = RP_ITEM_HEIGHT

        # Col B: Item code
        rpt_ws.cell(row=r, column=2, value=item_code).border = _thin_border()
        rpt_ws.cell(row=r, column=2).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

        # Col C: Description
        rpt_ws.cell(row=r, column=3, value=desc_val).border = _thin_border()
        rpt_ws.cell(row=r, column=3).alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

        # Col D: Category
        rpt_ws.cell(row=r, column=4, value=cat_val).border = _thin_border()
        rpt_ws.cell(row=r, column=4).alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

        # Col A: Image border
        rpt_ws.cell(row=r, column=1).border = _thin_border()

        # A: Product image
        safe_code = str(item_code).replace("/", "-").replace("\\", "-")
        img_path  = os.path.join(IMAGES_DIR, f"{safe_code}.png")
        if os.path.exists(img_path):
            try:
                img = XLImage(img_path)
                
                # Maintain aspect ratio to fit inside 135x135
                max_dim = 135.0
                ratio = min(max_dim / img.width, max_dim / img.height)
                new_w = int(img.width * ratio)
                new_h = int(img.height * ratio)
                img.width  = new_w
                img.height = new_h
                
                # Calculate EMU offset to center the image (1px = 9525 EMU)
                # Base margin is 5px (47625 EMU) + halfway of the remaining gap
                x_offset = 47625 + int((max_dim - new_w) / 2 * 9525)
                y_offset = 47625 + int((max_dim - new_h) / 2 * 9525)

                _from = AnchorMarker(col=0, colOff=x_offset, row=r-1, rowOff=y_offset)
                to    = AnchorMarker(col=1, row=r)
                img.anchor = TwoCellAnchor(editAs="oneCell", _from=_from, to=to)
                rpt_ws.add_image(img)
            except Exception:
                pass

    rpt_wb.save(out_xlsx)
    print(f"  Report saved: {out_xlsx}")

    # ── Export to PDF via win32com ──
    export_to_pdf_win32(out_xlsx, out_pdf)

    return out_xlsx, out_pdf


def export_to_pdf_win32(xlsx_path: str, pdf_path: str):
    """
    Uses win32com to open the xlsx in Excel and export it as PDF.
    Requires Microsoft Excel to be installed.
    """
    try:
        import win32com.client as win32
        abs_xlsx = str(Path(xlsx_path).resolve())
        abs_pdf  = str(Path(pdf_path).resolve())

        print(f"  Exporting to PDF: {abs_pdf}")
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(abs_xlsx)
        ws = wb.Worksheets(1)

        # Set print area to columns A:D
        ws.PageSetup.PrintArea = "A:D"
        ws.PageSetup.Orientation = 1   # xlPortrait
        
        # VERY IMPORTANT: To make FitToPagesWide work, Zoom must be False
        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = False

        # Set balanced margins (0.2 inches)
        margin_in = 0.2
        ws.PageSetup.LeftMargin   = excel.InchesToPoints(margin_in)
        ws.PageSetup.RightMargin  = excel.InchesToPoints(margin_in)
        ws.PageSetup.TopMargin    = excel.InchesToPoints(margin_in)
        ws.PageSetup.BottomMargin = excel.InchesToPoints(margin_in)
        ws.PageSetup.HeaderMargin = excel.InchesToPoints(0.1)
        ws.PageSetup.FooterMargin = excel.InchesToPoints(0.1)
        
        # Center on page
        ws.PageSetup.CenterHorizontally = True
        ws.PageSetup.CenterVertically = False

        wb.ExportAsFixedFormat(0, abs_pdf)  # 0 = xlTypePDF
        wb.Close(False)
        excel.Quit()
        print(f"  PDF exported: {abs_pdf}")

    except ImportError:
        print("  win32com not available. Skipping PDF export.")
        print("  To enable PDF export, run: pip install pywin32")
    except Exception as e:
        print(f"  PDF export error: {e}")

"""
excel_utils.py
Utility functions for Golden Inventory automation.
"""

import re
from datetime import datetime
from openpyxl import load_workbook
from scripts.item_overrides import ITEM_CATEGORY_OVERRIDES


MASTER_TAXONOMY_RULES = {
    "Major Appliances": {
        "Refrigerators": ["refrigerator", "fridge"],
        "Compact Refrigerators": ["compact refrigerator", "mini fridge"],
        "Beverage Centers": ["beverage center", "beverage cooler"],
        "Wine Coolers": ["wine cooler"],
        "Chest Freezers": ["chest freezer"],
        "Upright Freezers": ["upright freezer"],
        "Washers": ["washer"],
        "Dryers": ["dryer"],
        "Laundry Centers": ["laundry center"],
        "Dishwashers": ["dishwasher"],
        "Ranges": ["range"],
        "Cooktops": ["cooktop"],
        "Wall Ovens": ["wall oven"],
        "Range Hoods": ["range hood"]
    },
    "Kitchen Appliances": {
        "Air Fryers": ["air fryer", "fryer"],
        "Toasters": ["toaster", "2-slice toaster", "4-slice toaster"],
        "Toaster Ovens": ["toaster oven"],
        "Electric Burners": ["electric burner", "hot plate", "double burner", "buffet burner", "countertop burner"],
        "Slow Cookers": ["slow cooker", "crock pot"],
        "Pressure Cookers": ["pressure cooker", "pressure canner"],
        "Rice Cookers": ["rice cooker"],
        "Coffee Makers": ["coffee maker", "drip coffee maker"],
        "Espresso Makers": ["espresso maker", "espresso machine", "moka pot"],
        "Blenders": ["blender"],
        "Personal Blenders": ["personal blender"],
        "Food Processors": ["food processor"],
        "Mixers": ["hand mixer", "stand mixer"],
        "Microwaves": ["microwave"]
    },
    "Air & Climate Control": {
        "Portable Air Conditioners": ["portable air conditioner", "portable ac"],
        "Window Air Conditioners": ["window air conditioner", "window ac"],
        "Tower Fans": ["tower fan"],
        "Box Fans": ["box fan"],
        "Fans": ["fan", "ceiling fan", "drum fan", "window fan", "portable fan"],
        "Air Purifiers": ["air purifier", "purifier"],
        "Humidifiers": ["humidifier", "ultrasonic humidifier"]
    },
    "Audio": {
        "Bluetooth Speakers": ["bluetooth speaker", "portable speaker", "speaker", "boombox"],
        "Party Speakers": ["party speaker", "karaoke system", "speaker system"],
        "Headphones": ["headphone", "headset", "earphone", "earbuds", "ear buds"],
        "Earphones": ["earphones", "earbuds", "ear buds", "in-ear"],
        "Radios": ["radio", "am/fm", "fm radio", "clock radio"]
    },
    "Electronics": {
        "TV Mounts": ["tv wall mount", "full motion mount", "tilt mount", "fixed mount"],
        "LED Lighting": ["led strip light", "led light strip", "led light", "lighting strip", "ring light"],
        "Remotes": ["remote", "remote control", "universal remote"],
        "CD Players": ["cd player", "portable cd", "compact disc player", "cd boombox"],
        "Timers": ["timer", "countdown timer"],
        "Computing": ["mouse", "keyboard", "media box"],
        "Antennas": ["antenna", "passive antenna", "hd antenna", "hdtv antenna"]
    },
    "Home & Laundry": {
        "Steam Irons": ["steam iron", "steam/dry iron", "dry steam iron"],
        "Garment Steamers": ["garment steamer", "clothes steamer"]
    },
    "Health & Personal Care": {
        "Hair Dryers": ["hair dryer"],
        "Curling Irons": ["curling iron"],
        "Flat Irons": ["flat iron"],
        "Trimmers": ["trimmer"],
        "Shavers": ["shaver"],
        "Massagers": ["massager"]
    },
    "Home": {
        "Alarm Clocks": ["alarm clock"],
        "Flashlights": ["flashlight"],
        "Lamps": ["lamp"],
        "Bathroom Scales": ["bathroom scale"],
        "Calculators": ["calculator"]
    }
}


# ─── OVERRIDE → PARENT MAPPING ───────────────────────────────────────────────
# Maps override sub-category names (from item_overrides.py) that do NOT exactly
# match a MASTER_TAXONOMY_RULES key to their correct parent category.
# This ensures every string override resolves to a proper (parent, sub) pair.

OVERRIDE_PARENT_MAP = {
    # Major Appliances
    "Dishwasher": "Major Appliances",
    "Beverage Cooler": "Major Appliances",
    "Wine Cooler": "Major Appliances",
    "Refrigerators": "Major Appliances",

    # Kitchen Appliances
    "Air Fryers": "Kitchen Appliances",
    "Toasters": "Kitchen Appliances",
    "Toaster Ovens": "Kitchen Appliances",
    "Electric Burners": "Kitchen Appliances",
    "Slow Cookers": "Kitchen Appliances",
    "Pressure Cookers": "Kitchen Appliances",
    "Rice Cookers": "Kitchen Appliances",
    "Coffee Makers": "Kitchen Appliances",
    "Espresso Makers": "Kitchen Appliances",
    "Blenders": "Kitchen Appliances",
    "Personal Blenders": "Kitchen Appliances",
    "Food Processors": "Kitchen Appliances",
    "Mixers": "Kitchen Appliances",
    "Microwaves": "Kitchen Appliances",

    # Air & Climate Control
    "Portable Air Conditioners": "Air & Climate Control",
    "Window Air Conditioners": "Air & Climate Control",
    "Tower Fans": "Air & Climate Control",
    "Box Fans": "Air & Climate Control",

    # Audio
    "Bluetooth Speakers": "Audio",
    "Party Speakers": "Audio",

    # Electronics
    "TV Mounts": "Electronics",
    "LED Lighting": "Electronics",

    # Home & Laundry
    "Steam Irons": "Home & Laundry",
    "Garment Steamers": "Home & Laundry",

    # Health & Personal Care
    "Hair Dryers": "Health & Personal Care",
    "Curling Irons": "Health & Personal Care",
    "Flat Irons": "Health & Personal Care",
    "Trimmers": "Health & Personal Care",
    "Shavers": "Health & Personal Care",
    "Massagers": "Health & Personal Care",

    # Home
    "Alarm Clocks": "Home",
    "Flashlights": "Home",
    "Lamps": "Home",
    "Bathroom Scales": "Home",
    "Calculators": "Home"
}


# ─── SKU NORMALIZATION ────────────────────────────────────────────────

def normalize_sku(sku: str) -> str:
    if not sku:
        return ""

    return (
        str(sku)
        .strip()
        .upper()
        .replace(" ", "")
        .replace("/", "-")
        .replace("--", "-")
    )

NORMALIZED_ITEM_CATEGORY_OVERRIDES = {normalize_sku(k): v for k, v in ITEM_CATEGORY_OVERRIDES.items()}


# ─── TAXONOMY ENGINE ────────────────────────────────────────────────

def get_taxonomy(product_category_path: str, item_id: str = None, description: str = "", condition: str = ""):
    sku = normalize_sku(item_id)

    # 1. Determine refurbished status 
    # Checking both the SKU suffix and the condition string
    is_refurbished = (
        sku.endswith(("/RBO", "/RB", "/ROA", "-RBO", "-RB", "-ROA")) or 
        "refurbished" in str(condition).lower()
    )

    # 2. Step: Explicit item override (highest priority)
    if sku in NORMALIZED_ITEM_CATEGORY_OVERRIDES:
        override = NORMALIZED_ITEM_CATEGORY_OVERRIDES[sku]
        
        # Handle dict vs string override format
        sub_cat = override.get("sub", "Uncategorized") if isinstance(override, dict) else override
        parent_cat = override.get("parent", "Other") if isinstance(override, dict) else OVERRIDE_PARENT_MAP.get(override, "Other")
        
        if is_refurbished:
            return {"parent": "Refurbished", "sub": sub_cat}
        return {"parent": parent_cat, "sub": sub_cat}

    # 3. Step: Keyword scoring
    test_str = " ".join([str(product_category_path or ""), str(description or ""), str(item_id or "")]).lower()
    
    best_parent = None
    best_sub = None
    best_score = 0
    
    # Keyword scoring logic
    for parent, subs in MASTER_TAXONOMY_RULES.items():
        for sub, keywords in subs.items():
            score = sum(1 for kw in keywords if kw in test_str)
            if score > best_score:
                best_score = score
                best_parent = parent
                best_sub = sub

    # 4. Return match with Refurbished Priority
    if best_parent:
        if is_refurbished:
            return {"parent": "Refurbished", "sub": best_sub}
        return {"parent": best_parent, "sub": best_sub}

    if is_refurbished:
        return {"parent": "Refurbished", "sub": "Uncategorized"}

    return {"parent": "Other", "sub": "Uncategorized"}


# ─── RAW DATA LOADER ────────────────────────────────────────────────

def load_raw_data(inventory_tool_path: str) -> list[dict]:
    wb = load_workbook(inventory_tool_path, data_only=True)

    if "RAW" not in wb.sheetnames:
        raise ValueError("RAW sheet missing")

    ws = wb["RAW"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() for h in rows[0]]

    data = []
    for row in rows[1:]:
        row_dict = dict(zip(headers, row))
        if any(v for v in row_dict.values() if v is not None):
            data.append(row_dict)

    wb.close()
    return data


# ─── CATEGORY WRAPPER ────────────────────────────────────────────────

def classify_category(product_category_path, item_id, description, condition=""):
    return get_taxonomy(product_category_path, item_id, description, condition=condition)["sub"]

def get_latest_inventory_sheet(workbook):
    """
    Returns the newest timestamped inventory sheet and its date.
    """

    latest_sheet = None
    latest_datetime = None

    for sheet in workbook.sheetnames:

        if sheet == "Item Listing and Pricing":
            continue

        try:
            sheet_dt = datetime.strptime(sheet, "%b %d %Y - %I-%M %p")

            if latest_datetime is None or sheet_dt > latest_datetime:
                latest_datetime = sheet_dt
                latest_sheet = sheet

        except ValueError:
            continue

    if latest_sheet is None:
        raise Exception("No timestamped inventory sheet found.")

    return latest_sheet, latest_datetime.date()

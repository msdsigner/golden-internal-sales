import os
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import openpyxl
import json
from datetime import datetime

from scripts.excel_utils import get_latest_inventory_sheet, normalize_sku

def sync_inventory():
    year = datetime.now().year
    preferred = os.path.join("output", f"{year} Golden Inventory Internal Sales.xlsx")
    legacy = os.path.join("output", f"{year} Golden Inventory.xlsx")
    excel_path = preferred if os.path.exists(preferred) else legacy
    json_path = os.path.join("webapp", "data", "inventory.json")

    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {preferred} or {legacy}")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    sheet_name, recent_date = get_latest_inventory_sheet(wb)

    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found")
        return

    print(f"Syncing data from Excel sheet: {sheet_name}")
    sheet = wb[sheet_name]

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        print("Error: Sheet is empty")
        return

    # ── FIND THE REAL HEADER ROW ─────────────────────────────
    header_idx = -1
    for i, row in enumerate(rows):
        if any(
            str(cell).strip().lower() in ["item", "internal reference"]
            for cell in row
            if cell
        ):
            header_idx = i
            break

    if header_idx == -1:
        raise ValueError("Could not locate inventory header row.")

    headers = [
        str(h).strip().lower() if h else ""
        for h in rows[header_idx]
    ]

    col = {
        "item": None,
        "price": None,
        "qty": None
    }

    for i, h in enumerate(headers):
        if h == "item" or "internal reference" in h:
            col["item"] = i
        elif "price" in h:
            col["price"] = i
        elif (
            "qty" in h
            or "quantity" in h
            or "available" in h
            or "on hand" in h
        ):
            col["qty"] = i

    if col["item"] is None:
        raise ValueError("Item/SKU column not found.")

    excel_data = {}

    for row in rows[header_idx + 1:]:
        sku = row[col["item"]]
        if not sku:
            continue

        price = row[col["price"]] if col["price"] is not None else 0
        qty = row[col["qty"]] if col["qty"] is not None else 0

        try:
            qty = int(float(qty))
        except:
            qty = 0

        sku_key = normalize_sku(sku)
        excel_data[sku_key] = {
            "price": str(price) if price is not None else "0.00",
            "qty": qty
        }

    # ── LOAD JSON
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    items = data.get("items", [])
    updates = 0
    missing_in_excel = []
    missing_in_json = []

    publish_pricing = data.get("publish_pricing", True)
    json_map = {}

    # ── BUILD JSON MAP
    for item in items:
        sku = normalize_sku(item.get("id", ""))
        json_map[sku] = item

    # ── UPDATE JSON FROM EXCEL
    for sku, values in excel_data.items():
        if sku not in json_map:
            missing_in_json.append(sku)
            continue

        item = json_map[sku]
        new_price = values["price"] if publish_pricing else "Hidden"
        new_qty = values["qty"] if publish_pricing else "Hidden"

        if str(item.get("price")) != str(new_price) or str(item.get("available")) != str(new_qty):
            item["price"] = new_price
            item["available"] = new_qty
            updates += 1

    # ── FIND JSON ITEMS NOT IN EXCEL
    for sku in json_map:
        if sku not in excel_data:
            missing_in_excel.append(sku)

    # ── ALWAYS USE SAME DATE SOURCE
    data["last_updated"] = recent_date.strftime("%B %d, %Y")

    # SAVE JSON
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)

    print("\nSync complete.")
    print(f"Updated items: {updates}")
    print(f"Missing in Excel: {len(missing_in_excel)}")
    print(f"Missing in JSON: {len(missing_in_json)}")

if __name__ == "__main__":
    sync_inventory()
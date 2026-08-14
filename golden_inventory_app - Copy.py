"""
golden_inventory_app.py

Golden Inventory Automation Engine

Workflow:
    sync_odoo.py
        ↓
    Creates latest Inventory Tool.xlsx
        ↓
    Reads ONLY the newest Inventory Tool
        ↓
    Updates the master workbook
        ↓
    Generates timestamped Excel/PDF reports
        ↓
    Uploads latest reports to Google Drive
        ↓
    Regenerates website JSON database

Output:
    output/
        [Year] Golden Inventory.xlsx               (Master workbook)
        Golden Inventory - MMM DD YYYY - HH-MM AM.xlsx
        Golden Inventory - MMM DD YYYY - HH-MM AM.pdf

Input:
    input/
        Latest Inventory Tool.xlsx
        [Year] Golden Inventory.xlsx (Template if master doesn't exist)
"""

import os
import glob
import shutil
import json
import subprocess
import webbrowser
import time
import datetime
import re
import sys
from datetime import date

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    from openpyxl.utils import get_column_letter
except Exception as e:
    print("[Error] Missing dependency 'openpyxl' or failed to import it.")
    print("Please install openpyxl: pip install openpyxl")
    import sys
    sys.exit(1)

from scripts.export_service import create_report_and_pdf

# ─── Configuration ────────────────────────────────────────────────────────────
INPUT_DIR       = "input"
OUTPUT_DIR      = "output"
IMAGES_DIR      = "images"
INVENTORY_TOOL_PATTERN = os.path.join(INPUT_DIR, "Inventory Tool*.xlsx")
BANNER_PATH     = os.path.join(IMAGES_DIR, "banner.png")

# Row/column constants for Item Listing and Pricing
ILP_HEADER_ROW  = 3
ILP_DATA_START  = 4
ILP_COL_IMAGE   = 1   # A
ILP_COL_ITEM    = 2   # B
ILP_COL_DESC    = 3   # C
ILP_COL_CAT     = 4   # D
ILP_COL_PRICE   = 5   # E
ILP_COL_AVAIL   = 6   # F

# Row heights (Points: 1px ≈ 0.75pt)
ROW_DATE_HEIGHT   = 18.75  # 25px
ROW_BANNER_HEIGHT = 120    # 160px
ROW_ITEM_HEIGHT   = 105    # 140px

# Column widths (approx characters)
COL_WIDTHS = {1: 20, 2: 18, 3: 40, 4: 22, 5: 12, 6: 10}

# Header fill color
HEADER_FILL_YELLOW = PatternFill("solid", fgColor="FFC000")
HEADER_FILL_DARK   = PatternFill("solid", fgColor="404040")
HEADER_FONT_WHITE  = Font(name="Calibri", bold=True, color="FFFFFF")
HEADER_FONT_BLACK  = Font(name="Calibri", bold=True, color="000000")

DATED_HEADERS = [
    "Internal Reference", "Name", "Category",
    "Product Category", "Sales Price", "Quantity On Hand"
]


# ─── Helpers ─────────────────────────────────────────────────────────────────

def find_inventory_tool_files():
    def natural_sort_key(s):
        # File without number comes first (treated as 0), then others in numeric order.
        match = re.search(r'Inventory Tool\s*(\d+)', s)
        if match:
            return (1, int(match.group(1)))
        return (0, 0)
        
    files = glob.glob(INVENTORY_TOOL_PATTERN)
    files.sort(key=natural_sort_key)
    if not files:
        raise FileNotFoundError(f"No Inventory Tool files found in {INPUT_DIR}/")
    return files


def get_golden_inventory_path(year: int) -> str:
    """Returns the path to the working Golden Inventory file in output/."""
    return os.path.join(OUTPUT_DIR, f"{year} Golden Inventory.xlsx")


def get_input_golden_inventory_path(year: int) -> str:
    """Returns the path to the reference Golden Inventory file in input/."""
    return os.path.join(INPUT_DIR, f"{year} Golden Inventory.xlsx")


# load_raw_data moved to scripts.excel_utils
from typing import Optional
from scripts.excel_utils import load_raw_data, classify_category


def image_path_for_item(item_code: str) -> Optional[str]:
    """Return path to item image if it exists, else None."""
    if not item_code:
        return None
    safe = item_code.replace("/", "-").replace("\\", "-")
    path = os.path.join(IMAGES_DIR, f"{safe}.png")
    return path if os.path.exists(path) else None


def build_vlookup_row(row: int, sheet_name: str) -> tuple:
    """Return (desc_formula, cat_formula, price_formula, avail_formula) for a given row."""
    sn = sheet_name.replace("'", "''")
    b = f"B{row}"
    desc  = f"=IFERROR(VLOOKUP({b},'{sn}'!A:B,2,0),\"UNCATEGORIZED\")"
    cat   = f"=IFERROR(VLOOKUP({b},'{sn}'!A:C,3,0),\"UNCATEGORIZED\")"
    price = f"=IFERROR(VLOOKUP({b},'{sn}'!A:E,5,0),\"SOLD OUT\")"
    avail = f"=IFERROR(VLOOKUP({b},'{sn}'!A:F,6,0),\"SOLD OUT\")"
    return desc, cat, price, avail


def style_thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ─── Sheet Writers ────────────────────────────────────────────────────────────

def write_dated_sheet(ws, raw_data: list[dict]):
    """Write headers and data rows to a new dated sheet."""
    # Headers
    for col_idx, header in enumerate(DATED_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E4057")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = style_thin_border()

    ws.row_dimensions[1].height = 30

    # Column widths
    col_widths_dated = {1: 20, 2: 40, 3: 22, 4: 45, 5: 12, 6: 18}
    for col, w in col_widths_dated.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Data
    for r_idx, row_dict in enumerate(raw_data, start=2):
        internal_ref  = row_dict.get("Internal Reference", "") or ""
        name          = row_dict.get("Name", "") or ""
        product_cat   = row_dict.get("Product Category", "") or ""
        sales_price   = row_dict.get("Sales Price") or ""
        qty_on_hand   = row_dict.get("Free To Use Quantity") or 0
        category      = classify_category(product_cat, internal_ref, name)

        values = [internal_ref, name, category, product_cat, sales_price, qty_on_hand]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = style_thin_border()

        ws.row_dimensions[r_idx].height = 20

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DATED_HEADERS))}{len(raw_data) + 1}"
    print(f"  Written {len(raw_data)} data rows to dated sheet.")


def update_item_listing_sheet(ws_ilp, new_sheet_name: str, item_codes: list[str], sheet_date: date):
    """
    Rebuild the Item Listing and Pricing sheet:
    - Row 1: Date header (Merged A1:F1)
    - Row 2: Banner (Merged A2:F2)
    - Row 3: Headers
    - Row 4+: B = item code, C–F = VLOOKUP formulas
    """
    # ── Set column widths ──
    for col, w in COL_WIDTHS.items():
        ws_ilp.column_dimensions[get_column_letter(col)].width = w

    # ── Row 1: Date ──
    ws_ilp.row_dimensions[1].height = ROW_DATE_HEIGHT
    ws_ilp.merge_cells("A1:F1")
    cell_date = ws_ilp["A1"]
    cell_date.value = sheet_date.strftime("%b %d %Y - %I:%M %p").replace(" 0", " ")
    cell_date.font = Font(bold=True, size=14, color="1F497D")
    cell_date.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 2: Banner Row ──
    ws_ilp.row_dimensions[2].height = ROW_BANNER_HEIGHT
    ws_ilp.merge_cells("A2:F2")

    # ── Header row 3 ──
    headers = ["IMAGE", "Item", "Description", "Category", "Price", "Avail"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_ilp.cell(row=ILP_HEADER_ROW, column=col_idx, value=header)
        cell.font = Font(bold=True, color="000000")
        if col_idx == 4:  # Category — yellow
            cell.fill = HEADER_FILL_YELLOW
        else:
            cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = style_thin_border()
    ws_ilp.row_dimensions[ILP_HEADER_ROW].height = 20

    # ── Data rows ──
    for r_idx, item_code in enumerate(item_codes, start=ILP_DATA_START):
        ws_ilp.row_dimensions[r_idx].height = ROW_ITEM_HEIGHT

        # B = item code
        ws_ilp.cell(row=r_idx, column=ILP_COL_ITEM, value=item_code)

        # C–F = VLOOKUP formulas
        desc, cat, price, avail = build_vlookup_row(r_idx, new_sheet_name)
        ws_ilp.cell(row=r_idx, column=ILP_COL_DESC, value=desc)
        ws_ilp.cell(row=r_idx, column=ILP_COL_CAT, value=cat)
        ws_ilp.cell(row=r_idx, column=ILP_COL_PRICE, value=price)
        ws_ilp.cell(row=r_idx, column=ILP_COL_AVAIL, value=avail)

        # A–F = borders and alignment
        # We start from 1 to include A (Image column) per user request
        for col_idx in range(1, ILP_COL_AVAIL + 1):
            cell = ws_ilp.cell(row=r_idx, column=col_idx)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = style_thin_border()

    # ── Auto-filter on row 3 ──
    ws_ilp.auto_filter.ref = f"A{ILP_HEADER_ROW}:{get_column_letter(ILP_COL_AVAIL)}{len(item_codes) + ILP_DATA_START - 1}"


def insert_images_into_ilp(ws_ilp, item_codes: list[str]):
    """Insert product images into column A, and banner into row 2."""
    # Banner
    if os.path.exists(BANNER_PATH):
        try:
            banner = XLImage(BANNER_PATH)
            banner.width  = 900
            banner.height = 160
            # Anchor to spanning A2:F2 (col_idx 0-5, to 6)
            _from = AnchorMarker(col=0, row=1)
            to    = AnchorMarker(col=6, row=2)
            banner.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=to)
            ws_ilp.add_image(banner)
            print("  Banner image inserted into A2.")
        except Exception as e:
            print(f"  Warning: could not insert banner: {e}")

    # Product images
    inserted = 0
    missing  = []
    for r_idx, item_code in enumerate(item_codes, start=ILP_DATA_START):
        img_path = image_path_for_item(item_code)
        if img_path:
            try:
                img = XLImage(img_path)
                
                # Maintain aspect ratio to fit inside 135x135
                max_dim = 135.0
                ratio = min(max_dim / img.width, max_dim / img.height)
                new_w = int(img.width * ratio)
                new_h = int(img.height * ratio)
                img.width  = new_w
                img.height = new_h
                
                # Calculate EMU offset to center the image (1px = 9525 EMU)
                x_offset = 47625 + int((max_dim - new_w) / 2 * 9525)
                y_offset = 47625 + int((max_dim - new_h) / 2 * 9525)

                col_idx = 0
                row_idx = r_idx - 1
                _from = AnchorMarker(col=col_idx, colOff=x_offset, row=row_idx, rowOff=y_offset)
                to    = AnchorMarker(col=col_idx + 1, row=row_idx + 1)
                img.anchor = TwoCellAnchor(editAs="oneCell", _from=_from, to=to)
                ws_ilp.add_image(img)
                inserted += 1
            except Exception as e:
                print(f"  Warning: could not insert image for {item_code}: {e}")
                missing.append(item_code)
        else:
            missing.append(item_code)

    print(f"  Product images inserted: {inserted}, missing: {len(missing)}")
    if missing:
        missing_path = "output/missing_images.txt"
        os.makedirs("output", exist_ok=True)
        with open(missing_path, "w") as f:
            f.write("\n".join(missing))
        print(f"  Items with missing images saved to: {missing_path}")


def watch_mode():
    """Monitor input folder for changes and trigger run() automatically."""
    print("\n" + "="*60)
    print("  GOLDEN AUTO-SYNC: WATCH MODE ACTIVE")
    print("  Monitoring 'input/' folder for new Excel data...")
    print("  (Minimize this window and work in Excel - I'll handle the rest!)")
    print("="*60)

    def get_latest_mtime():
        files = glob.glob(INVENTORY_TOOL_PATTERN)
        if not files: return 0
        return max(os.path.getmtime(f) for f in files)

    last_mtime = get_latest_mtime()
    
    while True:
        try:
            time.sleep(5)
            current_mtime = get_latest_mtime()
            if current_mtime > last_mtime:
                print(f"\n[{datetime.datetime.now().strftime('%H:%M:%S')}] Detected change in Excel data. Syncing...")
                time.sleep(2)  # Give the OS a moment to finish the file write
                run(interactive=False)
                last_mtime = get_latest_mtime()
                print("\n" + "="*60)
                print("  SYNC COMPLETE. Watching for next change...")
                print("="*60)
        except KeyboardInterrupt:
            print("\nWatch mode stopped by user.")
            break
        except Exception as e:
            print(f"\n[Watcher Error] {e}")
            time.sleep(10)


# ─── Main Orchestrator ────────────────────────────────────────────────────────

def run(interactive=True):
    publish_pricing = True
    if interactive:
        ans = input("\n[Config] Publish Price and Quantity data to the website? [Y/n]: ").strip().lower()
        if ans == 'n':
            publish_pricing = False
            
    os.makedirs("webapp/data", exist_ok=True)
    with open("webapp/data/publish_config.json", "w") as f:
        json.dump({"publish_pricing": publish_pricing}, f)

    # 1. Find the newest Inventory Tool file only
    all_files = find_inventory_tool_files()

    if not all_files:
        raise Exception("No Inventory Tool files found.")

    latest_file = max(all_files, key=os.path.getmtime)

    tool_files = [latest_file]

    print(f"\n[1] Using latest Inventory Tool:")
    print(f"    {os.path.basename(latest_file)}")
        
    exported_reports = []

    # 2. Determine master workbook year (using latest file)
    mtime = os.path.getmtime(tool_files[-1])
    file_date = datetime.date.fromtimestamp(mtime)
    year = file_date.year
    gi_path = get_golden_inventory_path(year)

    # 3. Open or create master workbook (always in output/)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if os.path.exists(gi_path):
        print(f"[4] Opening existing output file: {gi_path}")
        wb = load_workbook(gi_path)
    else:
        # Try to seed from input/ reference copy first
        input_gi_path = get_input_golden_inventory_path(year)
        if os.path.exists(input_gi_path):
            print(f"[4] Seeding from input reference: {input_gi_path}")
            shutil.copy2(input_gi_path, gi_path)
            wb = load_workbook(gi_path)
        else:
            print(f"[4] Creating new workbook: {gi_path}")
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            wb.create_sheet("Item Listing and Pricing", 0)

    # 4. Process latest inventory snapshot

    tool_file = tool_files[0]

    print(f"\nProcessing Tool File: {tool_file}")

    # Use one timestamp for this entire run
    run_timestamp = datetime.datetime.now()
  
    curr_date = run_timestamp
    curr_sheet_name = curr_date.strftime("%b %d %Y - %I-%M %p").replace(" 0", " ")

    print(f"    Target Sheet: {curr_sheet_name}")

    # Load raw data
    raw_data = load_raw_data(tool_file)

    # Sort by category then SKU
    raw_data.sort(
        key=lambda x: (
            classify_category(
                str(x.get("Product Category", "")),
                str(x.get("Internal Reference", "")),
                str(x.get("Name", ""))
            ),
            str(x.get("Internal Reference", ""))
        )
    )

    print(f"    Loaded {len(raw_data)} product rows.")

    # Create or overwrite the snapshot sheet
    if curr_sheet_name in wb.sheetnames:
        print(f"    Updating existing sheet '{curr_sheet_name}'...")
        ws_dated = wb[curr_sheet_name]
        ws_dated.delete_rows(2, ws_dated.max_row)
    else:
        ws_dated = wb.create_sheet(curr_sheet_name, 1)
        print(f"    Created sheet '{curr_sheet_name}'.")

    # Build resolved data map
    resolved_data_map = {}

    for row_dict in raw_data:
        ref = str(row_dict.get("Internal Reference", "")).strip()

        if ref:
            resolved_data_map[ref] = {
                "desc": row_dict.get("Name", ""),
                "cat": classify_category(
                    row_dict.get("Product Category", ""),
                    item_id=ref,
                    description=row_dict.get("Name", "")
                )
            }

    write_dated_sheet(ws_dated, raw_data)

    # Rebuild Item Listing
    print("    Rebuilding Item Listing and Pricing...")

    ws_ilp = wb["Item Listing and Pricing"]

    if "A1:F1" in [str(m) for m in ws_ilp.merged_cells.ranges]:
        ws_ilp.unmerge_cells("A1:F1")

    if "A2:F2" in [str(m) for m in ws_ilp.merged_cells.ranges]:
        ws_ilp.unmerge_cells("A2:F2")

    for row in range(1, ws_ilp.max_row + 1):
        for col in range(1, 7):
            ws_ilp.cell(row=row, column=col).value = None

    ws_ilp._images = []

    item_codes = [
        str(r[0]).strip()
        for r in ws_dated.iter_rows(
            min_row=2,
            min_col=1,
            max_col=1,
            values_only=True
        )
        if r[0]
    ]

    update_item_listing_sheet(
        ws_ilp,
        curr_sheet_name,
        item_codes,
        curr_date
    )

    insert_images_into_ilp(ws_ilp, item_codes)

    if wb.sheetnames[0] != "Item Listing and Pricing":
        wb.move_sheet(
            "Item Listing and Pricing",
            offset=-wb.sheetnames.index("Item Listing and Pricing")
        )

    wb.save(gi_path)

    print(f"Saved: {gi_path}")

    print("Creating report...")

    xlsx_path, pdf_path = create_report_and_pdf(
        gi_path,
        curr_sheet_name,
        curr_date,
        year,
        data_map=resolved_data_map
    )

    exported_reports.append((xlsx_path, pdf_path))

    xlsx_id, pdf_id = None, None
    if exported_reports:
        recent_xlsx, recent_pdf = exported_reports[-1]
        print(f"\n[Drive] Uploading most recent reports to Google Drive...")
        try:
            from scripts.google_drive_upload import get_service, replace_file_on_gdrive
            FOLDER_ID = "1RbYL5fmeL0MhCC-E8CdpXMTEzyBiJH8x"
            service = get_service()
            if service:
                xlsx_id = replace_file_on_gdrive(service, recent_xlsx, FOLDER_ID, custom_name="Golden_Inventory_Latest.xlsx")
                pdf_id = replace_file_on_gdrive(service, recent_pdf, FOLDER_ID, custom_name="Golden_Inventory_Latest.pdf")
                print("[Drive] Successfully updated most recent snapshots on Drive.")
            else:
                print("[Drive] Authentication failed. Please ensure credentials.json is present.")
        except Exception as e:
            print(f"[Drive] Warning: Could not complete upload: {e}")

    try:
        from scripts.export_json import export_to_json
        print(f"\n[App] Regenerating Web App JSON Database...")
        export_to_json(xlsx_id=xlsx_id, pdf_id=pdf_id)
        
        # ─── AUTOMATIC CATEGORIZATION AUDIT ───
        print("\n" + "=" * 60)
        print("[Audit] Scanning for categorization accuracy...")
        with open("webapp/data/inventory.json", "r", encoding="utf-8") as f:
            data = json.load(f)
            all_items = data['items']
            uncat = [i['id'] for i in data['items'] if i['parent_category'] == 'Other' or i['sub_category'] == 'Uncategorized']
            # Save full taxonomy audit
            with open("categorized_items.txt", "w", encoding="utf-8") as f:
                f.write("ALL CATEGORIZED ITEMS AUDIT\n")
                f.write("=" * 60 + "\n\n")

                for item in sorted(all_items, key=lambda x: x['id']):
                    f.write(
                        f"SKU: {item['id']}\n"
                        f"Name: {item.get('name', '')}\n"
                        f"Category Field: {item.get('category', '')}\n"
                        f"Parent: {item.get('parent_category', '')}\n"
                        f"Sub: {item.get('sub_category', '')}\n"
                        f"{'-'*60}\n"
                    )
            if uncat:
                print(f"[ALERT] FOUND {len(uncat)} UNCATEGORIZED ITEMS: {', '.join(uncat[:10])}...")
                
                # Save complete list for review
                with open("uncategorized_items.txt", "w", encoding="utf-8") as f:
                    f.write("UNCATEGORIZED ITEMS\n")
                    f.write("=" * 60 + "\n\n")
                    for item in data['items']:
                        if item['id'] in uncat:
                            f.write(
                                f"{item['id']} | {item.get('name', '')}\n"
                            )

                print("[Audit] Full list saved to: uncategorized_items.txt")
                print("[Action] Please update 'scripts/excel_utils.py' or 'scripts/item_overrides.py' to fix these.")
            else:
                print("[Success] All items have been logically categorized.")
        
    except Exception as e:
        print(f"\n[App] Warning: Could not regenerate Web App JSON: {e}")

    # ── NEW: GitHub Pages Auto-Deploy ──
    try:
        if os.path.exists("github_config.json"):
            with open("github_config.json", "r") as f:
                g_config = json.load(f)
            if g_config.get("auto_deploy"):
                from scripts.deploy_to_github import deploy as github_deploy
                github_deploy()
    except Exception as e:
        print(f"\n[GitHub] Warning: Auto-deploy failed: {e}")

    # ── NEW: Unified Automation Sequence (Steps 2-4) ──
    try:
        # Step 2: High-level Sync
        from scripts.sync_inventory import sync_inventory
        print("\n" + "="*60)
        print("[Step 2/4] Synchronizing Web App Database...")
        sync_inventory()
        
        # Step 3: Deep Parity Audit & Auto-Fix
        from scripts.ensure_parity import ensure_parity
        print("\n" + "="*60)
        print("[Step 3/4] Running Parity Audit & Auto-Fix...")
        ensure_parity()
        
        # Step 4: System Health Check & Logging
        from scripts.validate_system import validate
        print("\n" + "="*60)
        print("[Step 4/4] Finalizing System Health Check...")
        validate()
        
    except Exception as e:
        print(f"\n[Error] Post-processing sequence failed: {e}")

    # Summary
    print("\n" + "=" * 60)
    print("  GOLDEN OPPORTUNITY CATALOG: UPDATE COMPLETE")
    print("=" * 60)

    if not interactive:
        print("\nAutomation complete. (Non-interactive mode)")
        return

    msg = input("\nWould you like to launch the local website preview? (y/n): ").strip().lower()
    if msg == 'y':
        print("\n[Preview] Starting local server at http://localhost:8000...")
        print("          (Close the browser and this window to stop the server)")
        webbrowser.open("http://localhost:8000")
        # Start the server (this is blocking, which is fine at the end of the app)
        subprocess.run(["python", "-m", "http.server", "8000", "--directory", "webapp"])
    else:
        print("\nAutomation complete. Goodbye!")


if __name__ == "__main__":
    if "--watch" in sys.argv:
        watch_mode()
    else:
        run(interactive=("--non-interactive" not in sys.argv))

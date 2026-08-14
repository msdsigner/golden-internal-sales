import json
import logging
import os
import re
import shutil
import sys
import xmlrpc.client
from pathlib import Path

from openpyxl import load_workbook


# --------------------------------------------------
# CONFIG
# --------------------------------------------------

ROOT = Path(__file__).resolve().parent

CONFIG_FILE = ROOT / "config_odoo.json"

with open(CONFIG_FILE, "r", encoding="utf-8") as f:
    CONFIG = json.load(f)

ODOO_URL = CONFIG["odoo"]["url"]
DB = CONFIG["odoo"]["database"]
USERNAME = CONFIG["odoo"]["username"]
API_KEY = CONFIG["odoo"]["api_key"]

WAREHOUSE_LOCATION_ID = CONFIG["inventory"]["warehouse_location_id"]
INCLUDED_TAG_IDS = CONFIG["inventory"]["included_tag_ids"]
MINIMUM_QTY = CONFIG["inventory"]["minimum_quantity"]

INPUT_FOLDER = ROOT / CONFIG["folders"]["input"]
RAW_SHEET = CONFIG["workbook"]["sheet"]


# --------------------------------------------------
# LOGGING
# --------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(message)s",
)

log = logging.getLogger(__name__)


# --------------------------------------------------
# INVENTORY TOOL VERSIONING
# --------------------------------------------------

def create_next_inventory_file():

    highest = 0

    pattern = re.compile(r"Inventory Tool(?: (\d+))?\.xlsx$", re.IGNORECASE)

    for file in INPUT_FOLDER.glob("Inventory Tool*.xlsx"):

        m = pattern.match(file.name)

        if not m:
            continue

        n = int(m.group(1)) if m.group(1) else 0

        highest = max(highest, n)

    source = (
        INPUT_FOLDER / "Inventory Tool.xlsx"
        if highest == 0
        else INPUT_FOLDER / f"Inventory Tool {highest}.xlsx"
    )

    destination = INPUT_FOLDER / f"Inventory Tool {highest+1}.xlsx"

    shutil.copy2(source, destination)

    log.info(f"Created {destination.name}")

    return destination


def sync_from_odoo():
    fields = [
        "brand_id",
        "default_code",
        "name",
        "free_qty",
        "categ_id",
        "product_tag_ids",
        "list_price",
        "barcode",
    ]

    print("Connecting to Odoo...")

    common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")

    uid = common.authenticate(
        DB,
        USERNAME,
        API_KEY,
        {}
    )

    if not uid:
        raise Exception("Authentication failed")

    models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object")

    print("Connected.")

    print("Finding products in Golden Opportunity warehouse...")

    quant_offset = 0
    limit = 500

    product_ids = set()
    warehouse_qty = {}

    while True:

        quants = models.execute_kw(
            DB,
            uid,
            API_KEY,
            "stock.quant",
            "search_read",
            [[
                ("location_id", "=", WAREHOUSE_LOCATION_ID),
                ("available_quantity", ">=", 1),
            ]],
            {
                "fields": [
                "product_id",
                "available_quantity",
                ],  
                "offset": quant_offset,
                "limit": limit,
            },
        )

        if not quants:
            break

        for q in quants:
            if not q["product_id"]:
                continue

            pid = q["product_id"][0]
            qty = q["available_quantity"]

            if qty <= 0:
                continue

            product_ids.add(pid)
            warehouse_qty[pid] = qty

        quant_offset += limit

    print(f"Found {len(product_ids)} products in warehouse.")

    products = []

    ids = list(product_ids)

    for i in range(0, len(ids), 500):

        batch_ids = ids[i:i + 500]

        batch = models.execute_kw(
            DB,
            uid,
            API_KEY,
            "product.product",
            "search_read",
            [[
                ("id", "in", batch_ids),
                ("product_tag_ids", "in", INCLUDED_TAG_IDS),
            ]],
            {
                "fields": fields,
            },
        )

        for p in batch:
            pid = p["id"]
            p["free_qty"] = warehouse_qty.get(pid, 0)
            products.append(p)

        print(f"Downloaded {len(products)} products...")

    # ---------------------------------------------------
    # Open Inventory Tool
    # ---------------------------------------------------

    inventory_file = create_next_inventory_file()

    wb = load_workbook(inventory_file)

    ws = wb[RAW_SHEET]

    # delete old rows
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    print("Writing Excel...")

    row = 2

    for p in products:

        brand = p["brand_id"][1] if p["brand_id"] else ""
        category = p["categ_id"][1] if p["categ_id"] else ""

        # For now don't resolve tags.
        # We'll optimize this later.
        tags = "New"

        ws.cell(row=row, column=1).value = brand
        ws.cell(row=row, column=2).value = p["default_code"]
        ws.cell(row=row, column=3).value = p["name"]
        ws.cell(row=row, column=4).value = p["free_qty"]
        ws.cell(row=row, column=5).value = category
        ws.cell(row=row, column=6).value = tags
        ws.cell(row=row, column=7).value = p["list_price"]
        ws.cell(row=row, column=8).value = p["barcode"]

        row += 1

        if row % 500 == 0:
            print(f"Wrote {row-2} rows...")

    wb.save(inventory_file)

    print("Done.")


if __name__ == "__main__":
    sync_from_odoo()